from pathlib import Path
import pandas as pd
import numpy as np
import json
from datetime import datetime
from openpyxl import load_workbook
import traceback
from typing import Union

DIR = Path(__file__).resolve().parent

# import PBL/own classes/modules
from ..general_functions import PRINT_COLORS
from ..IMAGE_tools.IMAGE_regions_settings import IMAGE_regions_nr2ISO
from ..IMAGE_tools import read_process_IMAGE_data as image_data
from ..historical_data import ISO, read_process_historical_data as hist_data
from . import log_downscaling_members as log

# global variables
IMAGE_START_YEAR = 1971
GWP = []
log_dataframes = []

output_dir = ""
current_dir = Path().cwd()
error_list = []

tmp_rows = pd.DataFrame(columns=["Value_sector", "Value_sum_individual", "ISO", "Year", "Sector", "Subsector"])

def _set_global_variables(Settings, GWP_CH4, GWP_N2O) -> None:
    global IMAGE_PROJECT, IMAGE_START_YEAR, HIST_YEAR_URBAN

    IMAGE_START_YEAR = Settings["IMAGE_START_YEAR"]
    HIST_YEAR = Settings["HIST_YEAR"]
    GWP.append(GWP_CH4) # CH4
    GWP.append(GWP_N2O) # N2O

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def _read_hist_variable(initiative_name: str, var_name_hist: str, hist_year:int, read_original:bool=True) -> pd.DataFrame:
    """
    initiative_name: Methane	Coal	Transport_road	Deforestation	Bunkers_aviation	BT_Power	BT_Road	Steel

    Description: reads in the historical data for the IMAGE variable based on the information in info
    pre: info csv file with TIMER to IEA translation
          and row i
    post: dataframe with historical data and the following columns names:  ["Name", "ISO3", "Country", "Year", "Value", "Unit"]
    """
    #  if file exists, read csv file
    var_name_hist_save = var_name_hist.replace("|", "_")
    hist_file = f"{DIR}/data/processed/df_hist_data_{var_name_hist_save}.csv"
    hist_path = Path(hist_file)
    if hist_path.exists() and not read_original:
        print(f"\nReading {str(hist_path)}")
        df_hist_data = pd.read_csv(hist_path, sep=";")
        hist_vars = df_hist_data["Variable"].unique()
        if var_name_hist in hist_vars:
            print(f"Variable {var_name_hist} already in {hist_path}")
            df_hist = df_hist_data[df_hist_data["Variable"]==var_name_hist]
        else:
            print(f"Appending {var_name_hist} to {hist_path}")
            df_hist = hist_data.ReadProcessHistoricalData(var_name_hist, hist_year)
            df_hist["Variable"] = var_name_hist
            df_hist_data = pd.concat([df_hist_data, df_hist], axis=0, ignore_index=True)
            df_hist_data.to_csv(hist_file, sep=";", index=False)
    else:
        # creating new {file hist_path}
        print(f"Creating new {hist_path}")
        df_hist_data = pd.DataFrame()
        df_hist = hist_data.ReadProcessHistoricalData(var_name_hist, hist_year)
        df_hist["Variable"] = var_name_hist
        df_hist_data = pd.concat([df_hist_data, df_hist], axis=0, ignore_index=True)
        df_hist_data.to_csv(hist_file, sep=";", index=False)

    # assumption: all files have the following columns: "ISO3", "Country", "Year", "Value"
    # if there is no World/WLD, this should be calculated
    countries_included = df_hist["Country"].unique()
    if "World" in countries_included:
        pass
    else:
        df_hist_world = df_hist.groupby(["Year"])[["Value"]].sum()
        df_hist_world.reset_index(inplace=True)
        df_hist_world["ISO3"] = "WLD"
        df_hist_world["Country"] = "World"
        print(f"World is not included in {initiative_name}, all countries are aggregated to world")
        df_hist = pd.concat([df_hist, df_hist_world], axis=0, ignore_index=True)

    return df_hist

def _read_process_hist_data(initiative_name: str, var_name_hist: str, hist_year:int, read_original:bool) -> tuple[pd.DataFrame, pd.DataFrame]:
    # read corresponding historical data for countries (all Year, all countries)
    df_hist = _read_hist_variable(initiative_name, var_name_hist, hist_year, read_original)
    print(f"Read historical data {var_name_hist} for {initiative_name}")

    # Calculate country share per region based on historical data
    # group historical data by IMAGE region and calculate aggregated value per region
    df_hist_IMAGE_Countries = ISO.compare_ISO_codes_for_hist_values(df_hist)

    # Aggregate country values to IMAGE regions
    df_hist_IMAGE_Regions = df_hist_IMAGE_Countries.groupby(by=["IMAGE_Region_Nr", "Year"])[["Value"]].sum()
    df_hist_IMAGE_Regions.reset_index(drop=False, inplace=True)
    df_hist_IMAGE_Regions = df_hist_IMAGE_Regions.loc[:,["IMAGE_Region_Nr", "Year", "Value"]]

    return (df_hist_IMAGE_Countries, df_hist_IMAGE_Regions)

def _calculate_country_share_in_region(df_hist_IMAGE_Countries: pd.DataFrame, hist_year: int, output_dir: str) -> pd.DataFrame:
    df_hist_IMAGE_Countries = df_hist_IMAGE_Countries.copy()[df_hist_IMAGE_Countries["Value"].notna()]
    var = df_hist_IMAGE_Countries["Variable"].iloc[0].replace("|", "_")

    # calculate % for country of region total from historical data
    mask= (df_hist_IMAGE_Countries["ISO_IMAGE"] == True) & (df_hist_IMAGE_Countries["ISO_hist"]==True) & (df_hist_IMAGE_Countries["Year"]==hist_year) & (~(df_hist_IMAGE_Countries["IMAGE_Region_Nr"]==28))
    df_var_country_share_in_region = df_hist_IMAGE_Countries.loc[mask].copy()
    df_var_country_share_in_region.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_var_country_share_in_region.csv", sep=";")
    df_var_country_share_in_region["perc_region"] = 100 * df_var_country_share_in_region["Value"] / df_var_country_share_in_region.groupby("IMAGE_Region_Nr")["Value"].transform("sum") # just for checking
    df_var_country_share_in_region= df_var_country_share_in_region.loc[:, ["ISO3", "Country_name", "IMAGE_Region_Nr", "Value"]]
    df_var_country_share_in_region["Country_share"] = df_var_country_share_in_region["Value"]/df_var_country_share_in_region.groupby("IMAGE_Region_Nr")["Value"].transform("sum")
    df_var_country_share_in_region.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_var_country_share_in_region_{var}.csv", sep=";")

    return df_var_country_share_in_region

def _read_initatives_country_shares(hist_year: int, input_file: str, output_dir, read_original:bool) -> pd.DataFrame:
    """
    xxx
    """
    # Read in paramters from json file
    try:
        parameters_individual_file = f"{DIR}/{input_file}"
        with open(parameters_individual_file) as f:
            Parameters = json.loads(f.read())
        settings = Parameters["Settings"]
        ICI_signatories_csv = settings["ICI_signatories_file"]
        init_members_scenario = settings["Members_Scenario"]
        df_initiative_members = pd.read_csv(f"{DIR}/data/input/{ICI_signatories_csv}", sep=";")
        ICI_scenario_parameters = Parameters["Init_Parameters_ICIs"]

    except ValueError as Err:
        print(f"\033[92mERROR in json file\033[0m\n")
        print(f"Run 'python -m json.tool {input_file}'")
        exit()

    # Read historical data for IMAGE variable
    df_initiatives_country_shares_collect = pd.DataFrame()
    for i, p in enumerate(ICI_scenario_parameters):
        if True:
            initiative_name = p["Initiative"]
            var_name_hist = p["Hist_variable"]
            print(f"\nRead in historical emissions {var_name_hist} for the {PRINT_COLORS["green"]}{initiative_name} initiative{PRINT_COLORS["end"]}")
            # read in historical data for initiative variable
            df_hist_IMAGE_Countries, df_hist_IMAGE_Regions = _read_process_hist_data(initiative_name, var_name_hist, hist_year, read_original)
            df_hist_IMAGE_Countries = df_hist_IMAGE_Countries[df_hist_IMAGE_Countries["Variable"].notna()]
            # Downscale results for ICI to members

            #df_hist_country_share_in_region = downscaler._calculate_country_share_in_region(df_hist_IMAGE_Countries, hist_year, output_dir)
            df_hist_country_share_in_region = _calculate_country_share_in_region(df_hist_IMAGE_Countries, hist_year, output_dir)

            # 1. calculate country shares in region
            print(f"Variable: {df_hist_IMAGE_Countries["Variable"].unique()}")
            # df_initiative_country_share = downscaler.calc_country_shares_in_region(df_hist_IMAGE_Countries, output_dir, 2020)
            # 2. Calculate share per region covered by members
            # df_initiative_members = downscaler.calculate_initiative_members_share_in_region(df_initiative_members, df_initiative_country_share, initiative_name, output_dir)

            if not (initiative_name=="Flaring"):
                share_collect = df_hist_country_share_in_region.copy()
                share_collect["ICI"] = initiative_name
                df_initiatives_country_shares_collect = pd.concat([df_initiatives_country_shares_collect, share_collect], axis=0, ignore_index=True)

    df_initiatives_country_shares_collect.rename(columns={"ISO3":"ISO", "IMAGE_Region_Nr":"IMAGE_region_nr"}, inplace=True)
    df_initiatives_country_shares_collect["Region"] = df_initiatives_country_shares_collect["IMAGE_region_nr"].map(IMAGE_regions_nr2ISO)
    df_initiatives_country_shares_collect = df_initiatives_country_shares_collect[~(df_initiatives_country_shares_collect["Country_share"].isnull())]
    df_initiatives_country_shares_collect.sort_values(by=["ICI", "ISO"], inplace=True)
    df_initiatives_country_shares_collect.drop(["Value"], axis=1, inplace=True)

    return df_initiatives_country_shares_collect

def _calc_reduction_elec(row):
# Reduction_country_individual_Coal
# Reduction_country_individual_Cooling
# Reduction_country_individual_Renewable
# Reduction_country_subsector_Coal
# Reduction_country_subsector_Cooling
# Reduction_country_subsector_Renewable

    match row["Reduction_criterium"]:
        case "reduction_subsector":
            reduction=max(row["Reduction_country_individual_Coal"], row["Reduction_country_individual_Renewable"], row["Reduction_country_individual_Cooling"])
        case "reduction_coal + reduction_cooling":
            reduction=row["Reduction_country_individual_Coal"] + row["Reduction_country_individual_Cooling"]
        case "reduction_coal":
            reduction=row["Reduction_country_individual_Coal"]
        case "COP28":
            reduction=row["Reduction_country_subsector"]
        case "reduction_cooling":
            reduction=row["Reduction_country_subsector"]
        case "baseline":
            reduction=0.0
        case _:
            print(f"\033[92mERROR: {row['Reduction_criterium']} is not a valid reduction criterium\033[0m")
            reduction=None

    return reduction

def _calculate_electricity_heat_reductions(df, output_dir):
    # Calculate the reductions in the electricity sector given the overlap between the initiatives that cover this sector
    # For each country and memberships of coal, renewable, and cooling --> calculate recuctions
    # Reduction_country_subsector --> reductions based on variable for individual initiatve (e.g. cooling, coal, renewable)
    # Reduction_country_individual --> reductions based on variable for sector (e.g. electricity)

    # If there are a few ducplicates in the data, they are removed
    duplicates = df.duplicated(["Region", "ISO", "Sector", "Subsector", "Year", "Unit", "ICI"])
    file_duplicates = f"{DIR}/runlog/downscale_members/{output_dir}/duplicates_electricity_heat.csv"
    # remove old duplicates file
    if Path(file_duplicates).exists():
        Path(file_duplicates).unlink()
    # if duplicates, save to runlog
    # CHECK --> there is only one initiative scenario now, so this produces many duplicates
    if duplicates.count()>0:
        print(f"WARNING: duplicates in the data, see {file_duplicates}")
        df_duplicates = df.loc[duplicates]
        df_duplicates.to_csv(file_duplicates, sep=";", index=False)
    df = df[~duplicates]

    # pivot to enable choosing the correct reduction (see downscaling/data/input/Electricity_reduction_parameters.xlsx)
    df = df.pivot(index=["Region", "Sector", "Subsector", "Year", "Unit", "ISO"], columns=["ICI"], values=["Reduction_country_individual", "Reduction_country_subsector"]).reset_index()
    #https://www.pauldesalvo.com/how-to-flatten-multiindex-columns-into-a-single-index-dataframe-in-pandas/
    df.columns = ['_'.join(col) for col in df.columns.values]
    df.columns = df.columns.map(lambda x: x[:-1] if x.endswith('_') else x)
    df = df.fillna(0)

    if "Reduction_country_individual_Coal" in df.columns:
        df["Membership_Coal"] = df["Reduction_country_individual_Coal"]!=0
    else:
        df["Reduction_country_individual_Coal"] = 0
        df["Reduction_country_subsector_Coal"] = 0
        df["Membership_Coal"] = False
    if "Reduction_country_individual_Renewable" in df.columns:
        df["Membership_Renewable"] =  df["Reduction_country_individual_Renewable"]!=0
    else:
        df["Reduction_country_individual_Renewable"] = 0
        df["Reduction_country_subsector_Renewable"] = 0
        df["Membership_Renewable"] = False
    if "Reduction_country_individual_Cooling" in df.columns:
        df["Membership_Cooling"] = df["Reduction_country_individual_Cooling"]!=0
    else:
        df["Reduction_country_individual_Cooling"] = 0
        df["Reduction_country_subsector_Cooling"] = 0
        df["Membership_Cooling"] = False

    conditions = [
        ((df["Membership_Coal"]==1) & (df["Membership_Renewable"]==1) & (df["Membership_Cooling"]==1)),
        ((df["Membership_Coal"]==1) & (df["Membership_Renewable"]==1) & (df["Membership_Cooling"]==0)),
        ((df["Membership_Coal"]==1) & (df["Membership_Renewable"]==0) & (df["Membership_Cooling"]==1)),
        ((df["Membership_Coal"]==1) & (df["Membership_Renewable"]==0) & (df["Membership_Cooling"]==0)),
        ((df["Membership_Coal"]==0) & (df["Membership_Renewable"]==1) & (df["Membership_Cooling"]==1)),
        ((df["Membership_Coal"]==0) & (df["Membership_Renewable"]==1) & (df["Membership_Cooling"]==0)),
        ((df["Membership_Coal"]==0) & (df["Membership_Renewable"]==0) & (df["Membership_Cooling"]==1)),
        ((df["Membership_Coal"]==0) & (df["Membership_Renewable"]==0) & (df["Membership_Cooling"]==0))
    ]
    #choices=[np.maximum(df["Reduction_country_subsector_Coal"], df["Reduction_country_subsector_Renewable"], df["Reduction_country_subsector_Cooling"]), 0, 0, 0, 0, 0, 0, 0]
    # https://datagy.io/pandas-conditional-column/
    choices=["reduction_subsector",
             "reduction_subsector",
             "reduction_coal + reduction_cooling",
             "reduction_coal",
             "COP28",
             "COP28",
             "reduction_cooling",
             "baseline"]

    # change subsector column names to the same name
    df["Reduction_country_subsector"] = df[["Reduction_country_subsector_Coal",
                                           "Reduction_country_subsector_Renewable",
                                           "Reduction_country_subsector_Cooling"]].max(axis=1, skipna=True)
    df["Reduction_criterium"] = np.select(conditions, choices, default=9999)
    df["Reduction_country_individual"] = df.apply(_calc_reduction_elec, axis=1)

    # drop columns that are not needed
    df.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_electricity_reductions.csv", sep=";", index=False)
    columns_to_drop = ["Reduction_country_subsector_Coal", "Reduction_country_subsector_Renewable", "Reduction_country_subsector_Cooling",
                       "Reduction_country_individual_Coal", "Reduction_country_individual_Renewable", "Reduction_country_individual_Cooling",
                       "Membership_Coal", "Membership_Renewable", "Membership_Cooling",
                       "Reduction_criterium"]
    df = df.drop(columns=[col for col in columns_to_drop if col in df.columns])
    # https://stackoverflow.com/questions/26886653/create-new-column-based-on-values-from-other-columns-apply-a-function-of-multi

    return df

def _calc_reduction_two_subsectors(row):
# Reduction_country_individual_Steel
# Reduction_country_individual_Cement
# Reduction_country_subsector_Steel
# Reduction_country_subsector_Cement

    global tmp_rows
    match row["Reduction_criterium"]:
        case "reduction_subsector":
            # exclude reductions due to other initiatives (as much as possble) by summing the individual country reductions
            new_row = pd.DataFrame({"ISO": row["ISO"], "Year": row["Year"],
                                    "Sector": row["Sector"], "Subsector": row["Subsector"],
                                    "Value_sector":[row["Reduction_country_subsector"]], "Value_sum_individual":[row["Reduction_country_individual_first"]+row["Reduction_country_individual_second"]]})
            if new_row.empty:
                print(f"\033[92mERROR: new_row is empty: {new_row}\033[0m")
            elif tmp_rows.empty:
                    tmp_rows = new_row
            else:
                tmp_rows = pd.concat([tmp_rows, new_row], axis=0, ignore_index=True)
            reduction=min(row["Reduction_country_subsector"], row["Reduction_country_individual_first"]+row["Reduction_country_individual_second"])
        case "reduction_first":
            reduction=row["Reduction_country_individual_first"]
        case "reduction_second":
            reduction=row["Reduction_country_individual_second"]
        case "baseline":
            reduction=0.0
        case _:
            print(f"\033[92mERROR: {row['Reduction_criterium']} is not a valid reduction criterium\033[0m")
            reduction=None

    return reduction

def _calculate_reductions_two_subsectors(df, output_dir):
    # Calculate the reductions in the assessed sector given the overlap between the initiatives that cover this sector
    # For each country and memberships of coal, renewable, and cooling --> calculate recuctions
    # Reduction_country_subsector --> reductions based on variable for individual initiatve (e.g. cooling, coal, renewable)
    # Reduction_country_individual --> reductions based on variable for sector (e.g. electricity)

    # check if two subsectors
    if len(df["ICI"].unique())!=2:
        print(f"\033[92mERROR: two subsectors are needed, but {len(df['Subsector'].unique())} are given\033[0m")
        df = pd.DataFrame()
        return df
    else:
        # If there are a few ducplicates in the data, they are removed
        duplicates = df.duplicated(["Region", "ISO", "Sector", "Subsector", "Year", "Unit", "ICI"])
        sector = df["Sector"].unique()[0]; subsector = df["Subsector"].unique()[0]
        file_duplicates = f"{DIR}/runlog/downscale_members/{output_dir}/duplicates_{sector}_{subsector}.csv"
        # remove old duplicates file
        if Path(file_duplicates).exists():
            Path(file_duplicates).unlink()
        # if duplicates, save to runlog
        # CHECK --> there is only one initiative scenario now, so this produces many duplicates
        if (duplicates == True).sum() > 0:
            print(f"WARNING: duplicates in the data, see {file_duplicates}")
            df_duplicates = df.loc[duplicates]
            df_duplicates.to_csv(file_duplicates, sep=";", index=False)
        df = df[~duplicates]

        # pivot to enable choosing the correct reduction (see downscaling/data/input/Electricity_reduction_parameters.xlsx)
        two_ICIs = df["ICI"].unique()
        df = df.pivot(index=["Region", "Sector", "Subsector", "Year", "Unit", "ISO"],
            columns=["ICI"],
            values=["Reduction_country_individual", "Reduction_country_subsector"]).reset_index()
        # Flatten the multi-level columns and add 'first' and 'second' as suffixes
        df.columns = ['_'.join(filter(None, col)).replace(' ', '').rstrip('_') for col in df.columns.values]
        #df.rename(columns=lambda x: x.replace("Reduction_country_individual_", "Reduction_country_individual_first_").replace("Reduction_country_subsector_", "Reduction_country_subsector_second_"), inplace=True)
        df.columns = df.columns.str.replace(f"_{two_ICIs[0]}", "_first")
        df.columns = df.columns.str.replace(f"_{two_ICIs[1]}", "_second")
        df = df.fillna(0)

        if "Reduction_country_individual_first" in df.columns:
            df["Membership_first"] = df["Reduction_country_individual_first"]!=0
        else:
            df["Reduction_country_individual_first"] = 0
            df["Reduction_country_subsector_first"] = 0
            df["Membership_first"] = False
        if "Reduction_country_individual_second" in df.columns:
            df["Membership_second"] =  df["Reduction_country_individual_second"]!=0
        else:
            df["Reduction_country_individual_second"] = 0
            df["Reduction_country_subsector_second"] = 0
            df["Membership_second"] = False

        conditions = [
            ((df["Membership_first"]==1) & (df["Membership_second"]==1) ),
            ((df["Membership_first"]==1) & (df["Membership_second"]==0) ),
            ((df["Membership_first"]==0) & (df["Membership_second"]==1) ),
            ((df["Membership_first"]==0) & (df["Membership_second"]==0) ),
        ]
        #choices=[np.maximum(df["Reduction_country_subsector_Coal"], df["Reduction_country_subsector_Renewable"], df["Reduction_country_subsector_Cooling"]), 0, 0, 0, 0, 0, 0, 0]
        # https://datagy.io/pandas-conditional-column/
        choices=["reduction_subsector",
                "reduction_first",
                "reduction_second",
                "baseline"]

        # change subsector column names to the same name
        df["Reduction_country_subsector"] = df[["Reduction_country_subsector_first", "Reduction_country_subsector_second"]].max(axis=1, skipna=True)
        df.drop(["Reduction_country_subsector_first", "Reduction_country_subsector_second"], axis=1, inplace=True)
        df["Reduction_criterium"] = np.select(conditions, choices, default=9999)
        df["Reduction_country_individual"] = df.apply(_calc_reduction_two_subsectors, axis=1)

        tmp_rows.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/tmp_rows.csv", sep=";", index=False)

        # drop columns that are not needed
        df.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_{subsector}_reductions.csv", sep=";", index=False)
        columns_to_drop = ["Reduction_country_subsector_first", "Reduction_country_subsector_second",
                           "Reduction_country_individual_first", "Reduction_country_individual_second",
                           "Membership_first", "Membership_second",
                           "Reduction_criterium"]
        df = df.drop(columns=[col for col in columns_to_drop if col in df.columns])
        # https://stackoverflow.com/questions/26886653/create-new-column-based-on-values-from-other-columns-apply-a-function-of-multi

        return df

def _calculate_subsector_pathway_reductions(sector, subsector, initiatives, df_member_pathway, output_dir) -> pd.DataFrame:
    print(f"\nCalculate pathway for sector: {sector}, subsector: {subsector}, initiatives: {initiatives}")
    match sector:
            case "Energy Supply":
                if subsector=="Electricity and Heat":
                    initiatives_electricity_heat = ["Coal", "Renewable", "Cooling"]
                    mask_electricity_heat = (df_member_pathway["ICI"].isin(initiatives_electricity_heat))
                    df_member_pathway_subsector = df_member_pathway[mask_electricity_heat]
                    count_elec_initiatives = df_member_pathway[df_member_pathway["ICI"].isin(initiatives_electricity_heat)]["ICI"].unique().size
                    if count_elec_initiatives > 1:
                        df_member_pathway_subsector = _calculate_electricity_heat_reductions(df_member_pathway_subsector, output_dir)
                    else:
                        df_member_pathway_subsector = df_member_pathway[df_member_pathway["ICI"].isin(initiatives_electricity_heat)]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_electricity_heat.csv", sep=";", index=False)
                elif subsector=="Fugitive|Oil Production":
                    initiatives_fugitive = ["Flaring"]
                    mask_fugitives = (df_member_pathway["ICI"].isin(initiatives_fugitive))
                    df_member_pathway_subsector = df_member_pathway[mask_fugitives]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_fugitive.csv", sep=";", index=False)
                else:
                    df_member_pathway_subsector = pd.DataFrame()
            case "Energy Demand":
                if subsector=="Total":
                    initiatives_demand = ["Efficiency"]
                    mask_demand = (df_member_pathway["ICI"].isin(initiatives_demand))
                    df_member_pathway_subsector = df_member_pathway[mask_demand]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_efficiency.csv", sep=";", index=False)
                else:
                    print(f"\nWARNING: {sector}|{subsector} is unkown Industry sub-sector")
                    df_member_pathway_subsector = pd.DataFrame()
            case "Industry":
                if subsector=="Total":
                    initiatives_industry = ["Steel", "Cement"]
                    mask_industry = (df_member_pathway["ICI"].isin(initiatives_industry))
                    df_member_pathway_subsector = df_member_pathway[mask_industry]
                    count_industry_initiatives = df_member_pathway[df_member_pathway["ICI"].isin(initiatives_industry)]["ICI"].unique().size
                    if count_industry_initiatives > 1:
                        df_member_pathway_subsector.rename(columns={"Reduction_country_individual_Steel": "Reduction_country_individual_first",
                                                                    "Reduction_country_individual_Cement": "Reduction_country_individual_second",
                                                                    "Reduction_country_subsector_Steel": "Reduction_country_subsector_first",
                                                                    "Reduction_country_subsector_Cement": "Reduction_country_subsector_second"}, inplace=True)
                        df_member_pathway_subsector = _calculate_reductions_two_subsectors(df_member_pathway_subsector, output_dir)
                    else:
                        df_member_pathway_subsector = df_member_pathway[df_member_pathway["ICI"].isin(initiatives_industry)]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_industry.csv", sep=";", index=False)
                else:
                    print(f"\nWARNING: {sector}|{subsector} is unkown Industry sub-sector")
                    df_member_pathway_subsector = pd.DataFrame()
            case "Transportation":
                if subsector=="Total":
                    initiatives_transport_road = ["Transport_cars_buses", "Transport_trucks"]
                    mask_transport_road = (df_member_pathway["ICI"].isin(initiatives_transport_road))
                    df_member_pathway_subsector = df_member_pathway[mask_transport_road]
                    count_transport_road_initiatives = df_member_pathway[df_member_pathway["ICI"].isin(initiatives_transport_road)]["ICI"].unique().size
                    if count_transport_road_initiatives > 1:
                        df_member_pathway_subsector.rename(columns={"Reduction_country_individual_Transport_cars_buses": "Reduction_country_individual_first",
                                                                    "Reduction_country_individual_Transport_trucks": "Reduction_country_individual_second",
                                                                    "Reduction_country_subsector_Transport_cars_buses": "Reduction_country_subsector_first",
                                                                    "Reduction_country_subsector_Transport_trucks": "Reduction_country_subsector_second"}, inplace=True)
                        df_member_pathway_subsector = _calculate_reductions_two_subsectors(df_member_pathway_subsector, output_dir)
                    else:
                        df_member_pathway_subsector = df_member_pathway[df_member_pathway["ICI"].isin(initiatives_transport_road)]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_transport_road.csv", sep=";", index=False)
                else:
                    print(f"\nWARNING: {sector}|{subsector} is unkown Transport road sub-sector")
                    df_member_pathway_subsector = pd.DataFrame()
            case "International Bunkers":
                if subsector=="Total":
                    initiatives_international_aviation = ["Bunkers_aviation"]
                    mask_int_aviation = (df_member_pathway["ICI"].isin(initiatives_international_aviation))
                    df_member_pathway_subsector = df_member_pathway[mask_int_aviation]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_aviation.csv", sep=";", index=False)
                else:
                    print(f"\nWARNING: {sector}|{subsector} is unkown International Bunkers sub-sector")
                    df_member_pathway_subsector = pd.DataFrame()
            case "Buildings":
                if subsector=="Total":
                    initiatives_buildings = ["Efficiency"]
                    mask_buildings = (df_member_pathway["ICI"].isin(initiatives_buildings))
                    df_member_pathway_subsector = df_member_pathway[mask_buildings]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_buildings_total.csv", sep=";", index=False)
                else:
                    print(f"\nWARNING: {sector}|{subsector} is unkown Buildings sub-sector")
                    df_member_pathway_subsector = pd.DataFrame()
            case "LULUCF":
                if subsector=="Deforestation":
                    initiatives_LULUCF = ["Deforestation"]
                    mask_LULUCF = (df_member_pathway["ICI"].isin(initiatives_LULUCF))
                    df_member_pathway_subsector = df_member_pathway[mask_LULUCF]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_LULUCF_total.csv", sep=";", index=False)
                else:
                    print(f"\nWARNING: {sector}|{subsector} is unkown LULUCF sub-sector")
                    df_member_pathway_subsector = pd.DataFrame()
            case "Total":
                if subsector=="Total" and "Methane" in initiatives: # Methane in initiatives
                    initiatives_total_methane = ["Methane"]
                    mask_total_methane = (df_member_pathway["ICI"].isin(initiatives_total_methane))
                    df_member_pathway_subsector = df_member_pathway[mask_total_methane]
                    df_member_pathway_subsector.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_reduction_total.csv", sep=";", index=False)
                else:
                    print(f"\nWARNING: {sector}|{subsector} is unkown Total sub-sector")
                    df_member_pathway_subsector = pd.DataFrame()
            case _:
                print("WARNING: {sector}|{subsector} is unkown subsector")
                df_member_pathway_subsector = pd.DataFrame()
                return df_member_pathway_subsector

    # print(f"Sector: {sector}, Subsector: {subsector}")
    df_member_pathway_subsector.rename(columns={"Reduction_country_individual": "Reduction_country_members"}, inplace=True)
    columns_to_drop = ["Reduction_country_subsector", "ICI"]
    df_member_pathway_subsector = df_member_pathway_subsector.drop(columns=[col for col in columns_to_drop if col in df_member_pathway_subsector.columns])
    df_member_pathway_subsector = df_member_pathway_subsector.groupby(["Sector", "Subsector", "Region", "Unit", "Year"])["Reduction_country_members"].sum().reset_index()

    return df_member_pathway_subsector

def _write_data_to_output(df: pd.DataFrame, sheet_name: str, file_name: str, file_type="xlsx") -> str:
    global log_dataframes
    file = file_name + "." + file_type

    # returns filename with extension
    file = file_name + "." + file_type
    #print(f"Sheet: {sheet_name}")
    # check length sheet name
    if len(sheet_name) > 31:
        print("sheet name is too long")
        sheet_name = sheet_name[:31]

    # save to file
    match file_type:
        case "xlsx":
            if not Path(file).is_file():
                df.to_excel(file, sheet_name=sheet_name, index=False)
            else:
                # check if sheet exists
                wb = load_workbook(file, read_only=True)
                if sheet_name in wb.sheetnames:
                    print(f"Sheet {sheet_name} already exists in {file}")
                    exit()
                else:
                    # write to existing Excel file
                    with pd.ExcelWriter(file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
        case "csv":
            df.to_csv(f"{DIR}/runlog/downscale_members/{output_dir}/{sheet_name}.csv", sep=";", index=False)
        case "pkl":
            df_with_name = {sheet_name: df}
            log_dataframes.append(df_with_name)

    return file

def calcualte_member_share(df_IMAGE_emissions: pd.DataFrame, ici_members, initiatives_country_shares, output_dir) -> pd.DataFrame:
        # make sure input variables are not changed
        df_member_pathway_cntr = df_IMAGE_emissions.copy()
        df_initiatives_country_shares = initiatives_country_shares.copy()
        df_ici_members = ici_members.copy()

        # Fugitive ICI does not have country members, exclude and World will be aggregated at the and of the calculations
        df_member_pathway_cntr = df_member_pathway_cntr[~(df_member_pathway_cntr["ICI"]=="Flaring")]
        df_member_pathway_cntr = df_member_pathway_cntr[(df_member_pathway_cntr["Region"]!="World")]
        # 3.1b merge data with membership
        df_member_pathway_cntr = pd.merge(df_member_pathway_cntr, df_ici_members, on=["Region", "ICI"], how="left")
        # 3.1c. merge data with initiatives' country shares
        df_member_pathway_cntr.drop(["Country_name"], axis=1, inplace=True)
        #df_initiatives_country_shares.drop(["Country_name"], axis=1, inplace=True)
        df_member_pathway_cntr = pd.merge(df_member_pathway_cntr, df_initiatives_country_shares, on=["ICI", "Region", "ISO", "IMAGE_region_nr"], how="left")
        # keep regional member shares
        mask_share = (df_member_pathway_cntr["Year"]==2020)
        df_initiatives_member_shares = df_member_pathway_cntr.loc[mask_share, ["ICI", "Region", "ISO", "Membership", "Year", "Country_share"]].copy()
        df_initiatives_member_shares["member_share_region"] = df_initiatives_member_shares["Membership"]*df_initiatives_member_shares["Country_share"]
        df_initiatives_member_shares = df_initiatives_member_shares.groupby(["ICI", "Region", "Year"]).agg({"member_share_region":"sum"}).reset_index()
        df_initiatives_member_shares = pd.merge(df_initiatives_member_shares, df_member_pathway_cntr[["ICI", "Region", "Year", "Value_baseline_individual"]].drop_duplicates(), on=["ICI", "Region", "Year"], how="left")

        # add world
        df_initiatives_member_shares_World = df_initiatives_member_shares.groupby(["ICI", "Year"]).apply(
            lambda g: pd.Series({
                'member_share_region': (g["Value_baseline_individual"] * g["member_share_region"]).sum() / g['Value_baseline_individual'].sum()})
                ).reset_index()
        df_initiatives_member_shares_World["Region"] = "World"
        df_initiatives_member_shares = pd.concat([df_initiatives_member_shares, df_initiatives_member_shares_World], ignore_index=True)
        df_initiatives_member_shares.drop(["Value_baseline_individual"], axis=1, inplace=True)
        df_initiatives_member_shares.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_initiatives_member_shares.csv", sep=";", index=False)

        return

def downscale_to_sector_emissions(input_file:str, dir_RT: str, initiative_scenarios: list[str], initiative_scenarios_additional:str, GWP_CH4:float, GWP_N2O:float, hist_year:int, read_original_hist: bool, log_type="xlsx", output_dir = "test") -> pd.DataFrame:
    '''
    Returns df_overview_initiatives that includes iniatives and corresponding variables for which data is created in csv files
    Downscale Kyoto GHG emisions to initiative members by
    - Disaggregating emissions in such a way that they cover the emissions from initiatives
    - Calculate pathway for Members that follow COP26, COP28 or COP26/COP28 emissions
    - Calulcate pathway for Non-members that follow CP_GI
    - Calculate pathways for emissions not covered by initiatives follow CP_GI
    - Reductions individual initiatives are based one of the two individual scenarios (e.g. COP26, COP28)
    - Reductions subsector are based on the subsector scenario (e.g. COP)

    '''
    # TO DO
    # 1. make this function able to use both one and two scenarios for initiative_scenarios
    # 2. make sure that you can add two variables for reductions in the same sector (e.g. CO2/N2O)

    # Init
    print("\n0. Initiatialisation")
    # log file for data
    now = datetime.now().strftime("%Y%m%d-%H%M-%S")
    log_name = f"{DIR}/runlog/downscale_members/{output_dir}/calculate_downscaled_sector_emissions_{now}"
    f_log_name = "" # init string for filename

    # check if output_dir exist, otherwise create them
    if not Path(f"{DIR}/data/output/downscale_members/{output_dir}/sector").exists():
        # create output_dir
        Path(f"{DIR}/data/output/downscale_members/{output_dir}/sector").mkdir(parents=True, exist_ok=True)
    if not Path(f"{DIR}/runlog/downscale_members/{output_dir}").exists():
        # create output_dir
        Path(f"{DIR}/runlog/downscale_members/{output_dir}").mkdir(parents=True, exist_ok=True)

    # 0. Initialize
    # 1. Collect Membership countries per initiative
    # 2. Calculate individual and subsector reductions for regions
    # 3. Read in country share (of emissions) per IMAGE region for each initiative
    # 4. Read in IMAGE sector emissions data for each ICI
    # 5. Calculate membership scenarios by assuming members follow COP(s) scenario's and non-member baseline scenario
    # 6. Calculate emissions pathways for Kyoto and subsector emissions

    try:
        # Note:
        # Each ICI can have emissions is one or more sectors and subsectors
        # Each subsector has one emissions variable that represents the total emissions in each sector (sector variables)
        # However, an initiative could only have impact on a part of the emissions that a subsector (reduction variables)

        # 0. Initialize
        # settings
        input_file_path = f"{DIR}/{input_file}"
        with open(input_file_path) as param_file:
            Parameters = json.loads(param_file.read())
        Settings = Parameters["Settings"]
        print(f"Settings: {Settings}")
        ICI_signatories_csv = Settings["ICI_signatories_file"]
        baseline_scenario = Settings["Baseline_Scenario"]
        vars_bunkers = Settings["Variable_Bunkers"]
        init_members_scenario = Settings["Members_Scenario"]
        ICI_scenario_parameters = Parameters["Init_Parameters_ICIs"]

        # check number of scenarios
        if len(initiative_scenarios) > 1:
            print(f"CHECK: initiative_scenarios must have only one scenario")

        _set_global_variables(Settings, GWP_CH4, GWP_N2O)

        # 1. Read in files
        print("\n1. Read in files")

        # 1.1 Membership
        ici_members = pd.read_csv(f"{DIR}/data/input/{ICI_signatories_csv}", sep=";")
        ici_members.drop(["Included by NCI?"], axis=1, inplace=True)
        ici_members = ici_members.melt(id_vars=["ISO", "IMAGE_region_nr", "Country_name"], var_name="ICI", value_name="Membership")
        ici_members["Region"] = ici_members["IMAGE_region_nr"].map(IMAGE_regions_nr2ISO)
        f_logname = _write_data_to_output(ici_members, "ici_members", log_name, log_type)

        # 1.2 ICI country shares per region
        df_initiatives_country_shares = _read_initatives_country_shares(hist_year, input_file, output_dir, read_original_hist)
        df_initiatives_country_shares.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_initiatives_country_shares.csv", sep=";", index=False)
        df_initiatives_country_shares.rename(columns={"ISO3":"ISO", "IMAGE_Region_Nr":"IMAGE_region_nr"}, inplace=True)
        f_logname = _write_data_to_output(df_initiatives_country_shares, "initiatives_country_shares", log_name, log_type)

        # 1.3. IMAGE subsector emissions data
        # 1.3.1 Make overview of sector, subsector, ICI, scenario's for which emissions are calculated and add to one Dataframe
        df_overview_initiatives = pd.DataFrame()
        index_coverage = 0
        index_calc_reductions = 0
        for i, p in enumerate(ICI_scenario_parameters):
            # create overview with sector, subsector, ICI, scenario's for which emissions are calculated
            initiative_name = p["Sector"]
            initiative_subsector = p["Subsector"]
            initiative_ICI = p["Initiative"]
            var_name_ghg_individual = p["GHG_individual_variable"]
            var_name_ghg_subsector = p["GHG_subsector_variable"]
            #for j, v in enumerate(var_name_ghg_individual):
            dict_initiatives = {index_calc_reductions: {'Variable_individual': var_name_ghg_individual[0],
                                                        'Variable_subsector': var_name_ghg_subsector[0],
                                                        'Sector' : initiative_name,
                                                        'Subsector' : initiative_subsector,
                                                        'ICI': initiative_ICI,
                                                        'Scenario_subsector': initiative_scenarios_additional, # scenario where all ICI reductions are combined into one run
                                                        'Scenario_individual': initiative_scenarios[0]}}
            index_calc_reductions += 1
            df_tmp_initiatives = pd.DataFrame(dict_initiatives).transpose()
            df_overview_initiatives = pd.concat([df_overview_initiatives, df_tmp_initiatives])

        if "International Bunkers" in df_overview_initiatives["Sector"].unique():
            var_bunkers_indivdiual = df_overview_initiatives.loc[df_overview_initiatives["Sector"]=="International Bunkers", "Variable_individual"].iloc[0]
            var_bunkers_sector = df_overview_initiatives.loc[df_overview_initiatives["Sector"]=="International Bunkers", "Variable_subsector"].iloc[0]

        # add GHG to overview
        df_overview_initiatives["GHG"] = df_overview_initiatives["Variable_subsector"].str.split("|", n=3).str[1]
        df_overview_initiatives.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_overview_initiatives.csv", sep=";", index=False)
        df_initiatives_individual = df_overview_initiatives[["ICI", "Variable_individual"]].drop_duplicates()
        df_initiatives_GHG = df_overview_initiatives[["Sector", "Subsector", "ICI", "Variable_subsector", "GHG"]].drop_duplicates()
        df_initiatives_GHG = df_initiatives_GHG.groupby(["Sector", "Subsector", "ICI"])["GHG"].agg(", ".join).reset_index()
        df_sectors_GHG = df_initiatives_GHG[["Sector", "Subsector", 'GHG']].drop_duplicates()
        print(f"{output_dir}-Sectors/subsectors: {df_sectors_GHG}")
        df_overview_initiatives_individual = df_overview_initiatives[["ICI", "Variable_individual", "Scenario_individual"]].sort_values(by=["ICI"])
        df_overview_initiatives_individual.rename(columns={"Variable_individual":"Variable", "Scenario_individual":"Scenario"}, inplace=True)
        df_overview_initiatives_subsector = df_overview_initiatives[["ICI", "Sector", "Subsector", "Variable_subsector", "Scenario_subsector"]].sort_values(by=["ICI", "Sector", "Subsector"])
        df_overview_initiatives_subsector.rename(columns={"Variable_subsector":"Variable", "Scenario_subsector":"Scenario"}, inplace=True)
        df_overview_subsectors = df_overview_initiatives[["Sector", "Subsector", "Variable_subsector", "Scenario_subsector"]].drop_duplicates().sort_values(by=["Sector", "Subsector"])
        df_overview_subsectors.rename(columns={"Variable_subsector":"Variable", "Scenario_subsector":"Scenario"}, inplace=True)
        f_logname = _write_data_to_output(df_overview_initiatives, "overview_initiatives", log_name, log_type)

        vars_individual = df_overview_initiatives["Variable_individual"].sort_values().tolist()
        print(f"\nvars_individual: {vars_individual}")
        f_logname = _write_data_to_output(pd.DataFrame(vars_individual), "vars_individual", log_name, log_type)
        vars_subsector = df_overview_initiatives["Variable_subsector"].sort_values().tolist()
        print(f"\nvars_subsector: {vars_subsector}")
        f_logname = _write_data_to_output(pd.DataFrame(vars_subsector), "vars_subsector", log_name, log_type)

        # 2 Calculate emissions reductions per ICI/sector/subsector for scenario's wikth individual ICIs and  subsector ICI's scenarios
        # Baseline                      | Mitigation
        # Value_subsector               | Value_subsector                        | Value_individual
        # df_IMAGE_baseline_subsector   | df_IMAGE_scenarios_individual_subsector | df_IMAGE_scenarios_individual
        # 2.1 Process baseline emissions for individual and subsector mitigation individual variables for each ICI/sector/subsector
        # 2.1a variables individual initiatives
        df_IMAGE_baseline_individual = image_data.read_vars_CO2eq([baseline_scenario], vars_individual, dir_RT, GWP)
        df_IMAGE_baseline_individual = pd.merge(df_IMAGE_baseline_individual, df_overview_initiatives_individual, on="Variable", how="left")
        df_IMAGE_baseline_individual.rename(columns={"Scenario_x":"Scenario_baseline"}, inplace=True)
        df_IMAGE_baseline_individual.drop(columns=["Scenario_y"], inplace=True)
        df_IMAGE_baseline_individual = df_IMAGE_baseline_individual[(df_IMAGE_baseline_individual["Region"]!="World")]
        # TO DO --> sum variables CO2/N2O for ICI
        df_IMAGE_baseline_individual.drop(["Model", "Variable"], axis=1, inplace=True)
        df_IMAGE_baseline_individual.rename(columns={"Value": "Value_baseline_individual"}, inplace=True)
        f_logname = _write_data_to_output(df_IMAGE_baseline_individual, "IMAGE_baseline_individual", log_name, log_type)
        # 2.1b variables subsector
        df_IMAGE_baseline_subsector = image_data.read_vars_CO2eq([baseline_scenario], vars_subsector, dir_RT, GWP)
        df_IMAGE_baseline_subsector = pd.merge(df_IMAGE_baseline_subsector, df_overview_subsectors, on="Variable", how="left")
        df_IMAGE_baseline_subsector.rename(columns={"Scenario_x":"Scenario_baseline"}, inplace=True)
        df_IMAGE_baseline_subsector.drop(columns=["Scenario_y"], inplace=True)
        df_IMAGE_baseline_subsector = df_IMAGE_baseline_subsector[(df_IMAGE_baseline_subsector["Region"]!="World")]
        # TO DO --> sum variables CO2/N2O for ICI
        df_IMAGE_baseline_subsector.drop(["Model", "Variable"], axis=1, inplace=True)
        df_IMAGE_baseline_subsector.rename(columns={"Value": "Value_baseline_subsector"}, inplace=True)
        f_logname = _write_data_to_output(df_IMAGE_baseline_subsector, "IMAGE_baseline_subsector", log_name, log_type)
        #df_IMAGE_baseline_subsector_ICI = pd.merge(df_IMAGE_baseline_subsector, df_overview_initiatives_subsector, on=["Sector", "Subsector"], how="left")

        # 2.2 (regions) Calculate total emissions reductions for individual and subsector mitigation reduction variables for each ICI/sector/subsector
        #       variables reduction (includes overlaping emissions, so do not add up to global total)
        # 2.2a variables individual
        df_IMAGE_individual = image_data.read_vars_CO2eq(initiative_scenarios, vars_individual, dir_RT, GWP)
        df_IMAGE_individual = pd.merge(df_IMAGE_individual, df_overview_initiatives_individual, on="Variable", how="left")
        df_IMAGE_individual.rename(columns={"Scenario_x":"Scenario_individual"}, inplace=True)
        df_IMAGE_individual.drop(columns=["Scenario_y"], inplace=True)
        df_IMAGE_individual = df_IMAGE_individual[(df_IMAGE_individual["Region"]!="World")]
        # TO DO --> sum variables CO2/N2O for ICI
        df_IMAGE_individual.drop(["Model", "Variable"], axis=1, inplace=True)
        df_IMAGE_individual.rename(columns={"Variable": "Variable_individual", "Value": "Value_individual"}, inplace=True)
        df_IMAGE_individual = pd.merge(df_IMAGE_baseline_individual, df_IMAGE_individual, on=["ICI", "Region", "Year", "Unit"], how="left")
        df_IMAGE_individual_reduction = df_IMAGE_individual.copy()
        df_IMAGE_individual_reduction["Reduction_region_individual"] = df_IMAGE_individual_reduction["Value_baseline_individual"] - df_IMAGE_individual_reduction["Value_individual"]
        df_IMAGE_individual_reduction.drop(["Value_baseline_individual", "Value_individual", "Scenario_baseline"], axis=1, inplace=True)
        f_logname = _write_data_to_output(df_IMAGE_individual_reduction, "IMAGE_individual_reduction", log_name, log_type)
        # 2.2b variables subsector
        df_IMAGE_subsector = image_data.read_vars_CO2eq([initiative_scenarios_additional], vars_subsector, dir_RT, GWP)
        df_IMAGE_subsector = pd.merge(df_IMAGE_subsector, df_overview_subsectors, on="Variable", how="left")
        df_IMAGE_subsector.rename(columns={"Scenario_x":"Scenario_subsector"}, inplace=True)
        df_IMAGE_subsector.drop(columns=["Scenario_y"], inplace=True)
        df_IMAGE_subsector = df_IMAGE_subsector[(df_IMAGE_subsector["Region"]!="World")]
        # TO DO --> sum variables CO2/N2O for ICI
        df_IMAGE_subsector.drop(["Model"], axis=1, inplace=True)
        df_IMAGE_subsector.rename(columns={"Value": "Value_subsector"}, inplace=True)
        df_IMAGE_subsector = pd.merge(df_IMAGE_baseline_subsector, df_IMAGE_subsector, on=["Sector", "Subsector", "Region", "Year", "Unit"], how="left")
        df_IMAGE_subsector_reduction = df_IMAGE_subsector.copy()
        df_IMAGE_subsector_reduction["Reduction_region_subsector"] = df_IMAGE_subsector_reduction["Value_baseline_subsector"] - df_IMAGE_subsector_reduction["Value_subsector"]
        df_IMAGE_subsector_reduction.drop(["Value_baseline_subsector", "Value_subsector", "Scenario_baseline"], axis=1, inplace=True)
        df_IMAGE_subsector_reduction = df_IMAGE_subsector_reduction[["Region", "Unit", "Year", "Sector", "Subsector", "Variable", "Scenario_subsector", "Reduction_region_subsector"]]
        f_logname = _write_data_to_output(df_IMAGE_subsector_reduction, "IMAGE_subsector_reduction", log_name, log_type)

        # 2.3 Merge individual and subsector reductions for each ICI/sector/subsector
        df_overview_ICI_sector_subsector = df_overview_initiatives[["Sector", "Subsector", "ICI"]].drop_duplicates()
        df_IMAGE_subsector_reduction_ICI = pd.merge(df_IMAGE_subsector_reduction, df_overview_ICI_sector_subsector, on=["Sector", "Subsector"], how="left")
        df_IMAGE_reduction = pd.merge(df_IMAGE_individual_reduction, df_IMAGE_subsector_reduction_ICI, on=["ICI", "Region", "Year", "Unit"], how="right")
        df_IMAGE_reduction = df_IMAGE_reduction[["ICI", "Sector", "Subsector", "Scenario_individual", "Scenario_subsector", "Region", "Unit", "Year", "Reduction_region_individual", "Reduction_region_subsector"]]
        f_logname = _write_data_to_output(df_IMAGE_reduction, "IMAGE_reduction", log_name, log_type)


        # 3. (countries) Collect country data for membership scenario's (reductions),
        #    necassary for calculating Kyoto/sectoral emissions pathways at regional level accounting for overlap
        #    Assumptions for overlap: for each sector/subsector with overlapping initiatives, the largest country reduction is used
        print ("\n2. Calculate member reduction pathways per sector/subsector and country based on subsector")
        # 3.1 First merge all needed data (emissions per ICI/sector, variables used for reductions, membership, and country shares)
        # 3.1a Init member pathway for each ICI/subsector
        #df_member_pathway_reduction_cntr = pd.pivot(df_IMAGE_reduction, index=["Sector", "Subsector", "ICI", "Variable", "Region", "Year", "Unit"], columns="Scenario", values="Reduction").reset_index()
        df_member_pathway_reduction_cntr = df_IMAGE_reduction.copy()
        # Fugitive ICI does not have country members, exclude and World will be aggregated at the and of the calculations
        df_member_pathway_reduction_cntr = df_member_pathway_reduction_cntr[~(df_member_pathway_reduction_cntr["ICI"]=="Flaring")]
        df_member_pathway_reduction_cntr = df_member_pathway_reduction_cntr[(df_member_pathway_reduction_cntr["Region"]!="World")]
        f_logname = _write_data_to_output(df_member_pathway_reduction_cntr, "member_pathway_red_cntr_a", log_name, log_type)
        # 3.1b merge data with membership
        df_member_pathway_reduction_cntr = pd.merge(df_member_pathway_reduction_cntr, ici_members, on=["Region", "ICI"], how="left")
        f_logname = _write_data_to_output(df_member_pathway_reduction_cntr, "member_pathway_red_cntr_b", log_name, log_type)
        # 3.1c. merge data with initiatives' country shares
        df_member_pathway_reduction_cntr.drop(["Country_name"], axis=1, inplace=True)
        df_initiatives_country_shares.drop(["Country_name"], axis=1, inplace=True)
        df_member_pathway_reduction_cntr = pd.merge(df_member_pathway_reduction_cntr, df_initiatives_country_shares, on=["ICI", "Region", "ISO", "IMAGE_region_nr"], how="left")
        # keep regional member shares
        # mask_share = (df_member_pathway_reduction_cntr["Year"]==2020)
        # check_initiatives_country_shares = df_member_pathway_reduction_cntr.loc[mask_share, ["ICI", "Region", "ISO", "Membership", "Year", "Country_share"]].copy()
        # check_initiatives_country_shares["member_share_region"] = check_initiatives_country_shares["Membership"]*check_initiatives_country_shares["Country_share"]
        # check_initiatives_country_shares = check_initiatives_country_shares.groupby(["ICI", "Region", "Year"]).agg({"member_share_region":"sum"}).reset_index()
        # check_initiatives_country_shares.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/check_initiatives_country_shares.csv", sep=";", index=False)
        f_logname = _write_data_to_output(df_member_pathway_reduction_cntr, "initiatives_country_shares", log_name, log_type)
        # calcualte reductions
        df_member_pathway_reduction_cntr["Reduction_country_individual"] = df_member_pathway_reduction_cntr["Reduction_region_individual"]*df_member_pathway_reduction_cntr["Membership"]*df_member_pathway_reduction_cntr["Country_share"]
        df_member_pathway_reduction_cntr["Reduction_country_subsector"] = df_member_pathway_reduction_cntr["Reduction_region_subsector"]*df_member_pathway_reduction_cntr["Membership"]*df_member_pathway_reduction_cntr["Country_share"]
        df_member_pathway_reduction_cntr.drop(["IMAGE_region_nr", "Membership", "Country_share"], axis=1, inplace=True)
        df_member_pathway_reduction_cntr = df_member_pathway_reduction_cntr[["Sector", "Subsector", "ICI", "Scenario_individual", "Scenario_subsector", "Region", "ISO", "Year", "Unit", "Reduction_region_individual", "Reduction_country_individual", "Reduction_region_subsector", "Reduction_country_subsector"]]
        f_logname = _write_data_to_output(df_member_pathway_reduction_cntr, "member_pathway_red_cntr_c", log_name, log_type)

        # 4 Member pahtways for individual ICIs (per GHG)
        # 4.1 Calculate reductions per region --> for each sector/subsector member pathway; NOT based on overlap assumptions
        df_member_pathway_reduction_individual = df_member_pathway_reduction_cntr[["ICI", "Scenario_individual", "Region", "ISO", "Year", "Unit", "Reduction_country_individual"]].copy()
        df_member_pathway_reduction_individual = df_member_pathway_reduction_individual.groupby(["ICI", "Scenario_individual", "Region", "Year", "Unit"], observed=False)["Reduction_country_individual"].sum().reset_index()
        df_member_pathway_reduction_individual.rename(columns={"Reduction_country_individual": "Reduction_country_members"}, inplace=True)
        df_member_pathway_individual = df_member_pathway_reduction_individual.copy()
        df_IMAGE_baseline_individual_excl_flaring = df_IMAGE_baseline_individual[~(df_IMAGE_baseline_individual["ICI"]=="Flaring")]
        df_member_pathway_individual = pd.merge(df_IMAGE_baseline_individual_excl_flaring, df_member_pathway_reduction_individual, on=["ICI", "Region", "Year", "Unit"], how="left")
        df_member_pathway_individual["Value"] = df_member_pathway_individual["Value_baseline_individual"] - df_member_pathway_individual["Reduction_country_members"]
        df_member_pathway_individual.drop(["Value_baseline_individual", "Reduction_country_members"], axis=1, inplace=True)
        df_member_pathway_individual["Scenario"] = init_members_scenario
        # add world
        df_member_pathway_individual_World = df_member_pathway_individual.groupby(["ICI", "Scenario", "Scenario_baseline", "Scenario_individual", "Year", "Unit"], observed=False).agg({"Value":"sum"}).reset_index()
        df_member_pathway_individual_World["Region"] = "World"
        df_member_pathway_individual = pd.concat([df_member_pathway_individual, df_member_pathway_individual_World], axis=0, ignore_index=True)
        df_member_pathway_individual = df_member_pathway_individual[["ICI", "Scenario", "Scenario_baseline", "Scenario_individual", "Region", "Year", "Value", "Unit"]]
        df_member_pathway_individual.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_individual.csv", sep=";", index=False)
        f_logname = _write_data_to_output(df_member_pathway_individual, "member_pathway_individual_excl_flaring", log_name, log_type)
        # add Flaring for World
        mask_flaring = (df_IMAGE_reduction["ICI"]=="Flaring") & (df_IMAGE_reduction["Sector"]=="Energy Supply") & (df_IMAGE_reduction["Subsector"]=="Fugitive|Oil Production")
        df_member_pathway_individual_flaring = df_IMAGE_reduction[mask_flaring].copy()
        df_member_pathway_individual_flaring.drop(["Sector", "Subsector", "Scenario_subsector", "Reduction_region_subsector"], axis=1, inplace=True)
        df_member_pathway_individual_flaring = df_member_pathway_individual_flaring.groupby(["Scenario_individual", "ICI", "Year", "Unit"], observed=False)["Reduction_region_individual"].sum().reset_index()
        members_Flaring = pd.read_csv(f"{DIR}/data/input/ICI_members_flaring.csv", sep=";")
        members_Flaring_perc = members_Flaring.loc[members_Flaring["Year"]==2020, "Value"].values[0]
        df_member_pathway_individual_flaring["Reduction_region_individual"] *= members_Flaring_perc # assume global percentage applies to all regions
        df_IMAGE_baseline_individual_flaring = df_IMAGE_baseline_individual[df_IMAGE_baseline_individual["ICI"]=="Flaring"]
        df_IMAGE_baseline_individual_flaring.rename(columns={"Value_baseline_individual":"Value"}, inplace=True)
        df_IMAGE_baseline_individual_flaring = df_IMAGE_baseline_individual_flaring.groupby(["Scenario_baseline", "ICI", "Year", "Unit"], observed=False)["Value"].sum().reset_index()
        df_member_pathway_individual_flaring = pd.merge(df_IMAGE_baseline_individual_flaring, df_member_pathway_individual_flaring, on=["ICI", "Year", "Unit"], how="left")
        df_member_pathway_individual_flaring["Value"] = df_member_pathway_individual_flaring["Value"] - df_member_pathway_individual_flaring["Reduction_region_individual"]
        df_member_pathway_individual_flaring["Region"] = "World"
        df_member_pathway_individual_flaring["Scenario"] = init_members_scenario
        df_member_pathway_individual_flaring.drop(["Reduction_region_individual"], axis=1, inplace=True)

        df_member_pathway_individual = pd.concat([df_member_pathway_individual, df_member_pathway_individual_flaring], axis=0, ignore_index=True)
        df_member_pathway_individual.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_individual.csv", sep=";", index=False)
        f_logname = _write_data_to_output(df_member_pathway_individual, "member_pathway_individual", log_name, log_type)

        # # Calculate member shares
        # df_tmp = calcualte_member_share(df_IMAGE_individual, ici_members, df_initiatives_country_shares, output_dir)

        # Overview of total emissions, 1) covered/no-covered by country signatories, 2) covered/not-covered by iniatives
        #                                                 |  covered by initiatives | not covered by initiatives | TOTAL
        # -----------------------------------------------------------------------------------------------------------------------------
        #                                                 |                         |                            |
        # Emissions category covered by initiatives       |                         |                            |
        # (mitigation)                                    |                         |                            |
        # -----------------------------------------------------------------------------------------------------------------------------
        # Emissions category not covered by initiatives
        # (baseline)                                      |                         |                            |
        # ------------------------------------------------------------------------------------------------------------------------------
        # TOTAL                                           |                         |                            | Total Kyoto emissions

        # 5 Member pathways including overlap for sectors/subsectors
        # Calculate reductions per region --> for each sector/subsector member pathway; based on overlap assumptions if more than one initiative covers a sector/subsector/ghg
        # 5.1 Calculate reductions per sector/subsector based on sector/subsector variables
        print("\n3. Calculate member pahtways for each sector/subsector based on sector/subsector variables")
        sectors_subsectors = df_overview_initiatives[["Sector", "Subsector"]].drop_duplicates()
        print(f"sectors_subsectors_GHG: {pd.DataFrame(sectors_subsectors)}")
        # Calculate reductions per sector/subsector
        df_member_pathway_reduction_reg = pd.DataFrame()
        for row in sectors_subsectors.itertuples(index=False):
            print(f"Input: {row[0]}, subsector: {row[1]}")
            if row[0]=="Energy Supply" and row[1]=="Fugitive|Oil Production":
                mask_flaring = (df_IMAGE_reduction["ICI"]=="Flaring") & (df_IMAGE_reduction["Sector"]=="Energy Supply") & (df_IMAGE_reduction["Subsector"]=="Fugitive|Oil Production")
                df_member_pathway_reduction_reg_subsector = df_IMAGE_reduction[mask_flaring].copy()
                df_member_pathway_reduction_reg_subsector.drop(["ICI", "Scenario_individual", "Scenario_subsector", "Reduction_region_subsector"], axis=1, inplace=True)
                members_Flaring = pd.read_csv(f"{DIR}/data/input/ICI_members_flaring.csv", sep=";")
                members_Flaring_perc = members_Flaring.loc[members_Flaring["Year"]==2020, "Value"].values[0]
                df_member_pathway_reduction_reg_subsector["Reduction_region_individual"] *= members_Flaring_perc # assume global percentage applies to all regions
                df_member_pathway_reduction_reg_subsector.rename(columns={"Reduction_region_individual": "Reduction_country_members"}, inplace=True)
            else:
                mask_initiatives =  (df_overview_initiatives["Sector"]==row[0]) & (df_overview_initiatives["Subsector"]==row[1])
                initiatives = df_overview_initiatives.loc[mask_initiatives, "ICI"].unique()
                df_member_pathway_reduction_reg_subsector = _calculate_subsector_pathway_reductions(row[0], row[1], initiatives, df_member_pathway_reduction_cntr, output_dir)
            # sum to variable level
            df_member_pathway_reduction_reg = pd.concat([df_member_pathway_reduction_reg, df_member_pathway_reduction_reg_subsector], axis=0, ignore_index=True)
        f_logname = _write_data_to_output(df_member_pathway_reduction_reg, "member_pathway_reduction_reg_a", log_name, log_type)
        df_member_pathway_reduction_reg = pd.merge(df_member_pathway_reduction_reg, df_IMAGE_subsector_reduction, on=["Sector", "Subsector", "Region", "Year", "Unit"], how="left")
        df_member_pathway_reduction_reg.drop(["Scenario_subsector"], axis=1, inplace=True)
        df_member_pathway_reduction_reg["Reduction_check_before"] = (df_member_pathway_reduction_reg["Reduction_country_members"] > df_member_pathway_reduction_reg["Reduction_region_subsector"])
        df_member_pathway_reduction_reg["Reduction_country_members"] = df_member_pathway_reduction_reg[["Reduction_country_members", "Reduction_region_subsector"]].min(axis=1, skipna=True)
        df_member_pathway_reduction_reg["Reduction_check_after"] = (df_member_pathway_reduction_reg["Reduction_country_members"] > df_member_pathway_reduction_reg["Reduction_region_subsector"])
        f_logname = _write_data_to_output(df_member_pathway_reduction_reg, "member_pathway_reduction_reg_b", log_name, log_type)
        df_member_pathway_reduction_reg.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reduction_reg.csv", sep=";", index=False)
        df_member_pathway_reduction_reg.drop(["Reduction_region_subsector", "Reduction_check_before", "Reduction_check_after"], axis=1, inplace=True)
        f_logname = _write_data_to_output(df_member_pathway_reduction_reg, "member_pathway_reduction_reg", log_name, log_type)

        # 5.2 Create member pathway --> substract reduction from subsector variable emissions
        vars_subsectors = df_overview_initiatives[["Sector", "Subsector", "Variable_subsector"]].drop_duplicates()
        vars_subsectors.rename(columns={"Variable_subsector":"Variable"}, inplace=True)
        df_IMAGE_baseline_subsector_incl_info = pd.merge(df_IMAGE_baseline_subsector, vars_subsectors, on=["Sector", "Subsector"], how="left")
        df_IMAGE_baseline_subsector_incl_info.drop(["Scenario_baseline"], axis=1, inplace=True)
        df_IMAGE_baseline_subsector_reg = df_IMAGE_baseline_subsector_incl_info.groupby(["Sector", "Subsector", "Variable", "Region", "Year", "Unit"], observed=True)["Value_baseline_subsector"].sum().reset_index()
        df_member_pathway_reg_subsector_GHG = pd.merge(df_IMAGE_baseline_subsector_reg, df_member_pathway_reduction_reg, on=["Sector", "Subsector", "Variable", "Region", "Year", "Unit"], how="left")
        df_member_pathway_reg_subsector_GHG["Value"] = df_member_pathway_reg_subsector_GHG["Value_baseline_subsector"] - df_member_pathway_reg_subsector_GHG["Reduction_country_members"]
        df_member_pathway_reg_subsector_GHG["Scenario"] = init_members_scenario
        df_member_pathway_reg_subsector_GHG = df_member_pathway_reg_subsector_GHG.loc[:, ["Sector", "Subsector", "Variable", "Scenario", "Region", "Year", "Value", "Unit"]]
        # check if emissions for member scenarios are not lower than COP scenario (can happen with negative emissions), Sector	Subsector	Scenario	Region	Year	Value	Unit
        df_member_pathway_reg_subsector_GHG = pd.merge(df_member_pathway_reg_subsector_GHG, df_IMAGE_subsector, on=["Sector", "Subsector", "Variable", "Region", "Year", "Unit"], how="left")
        df_member_pathway_reg_subsector_GHG.rename(columns={"Scenario_x":"Scenario"}, inplace=True)
        df_member_pathway_reg_subsector_GHG["Change_value"] = (df_member_pathway_reg_subsector_GHG["Value"] < df_member_pathway_reg_subsector_GHG["Value_subsector"])
        df_member_pathway_reg_subsector_GHG["Value"] = df_member_pathway_reg_subsector_GHG[["Value", "Value_subsector"]].max(axis=1)
        df_member_pathway_reg_subsector_GHG = df_member_pathway_reg_subsector_GHG[["Sector", "Subsector", "Variable", "Scenario", "Region", "Year", "Value", "Unit"]]
        # add world
        df_member_pathway_reg_subsector_GHG_World = df_member_pathway_reg_subsector_GHG.groupby(["Sector", "Subsector", "Variable", "Scenario", "Year", "Unit"])["Value"].sum().reset_index() # "Sector", "Subsector",
        df_member_pathway_reg_subsector_GHG_World["Region"] = "World"
        df_member_pathway_reg_subsector_GHG = pd.concat([df_member_pathway_reg_subsector_GHG, df_member_pathway_reg_subsector_GHG_World], axis=0, ignore_index=True)
        # if sector is International Bunkers, then exclude regions except World
        if "International Bunkers" in df_member_pathway_reg_subsector_GHG["Sector"].unique():
            mask_bunkers = (df_member_pathway_reg_subsector_GHG["Sector"]=="International Bunkers") & ~(df_member_pathway_reg_subsector_GHG["Region"]=="World")
            df_member_pathway_reg_subsector_GHG = df_member_pathway_reg_subsector_GHG[~mask_bunkers]
        df_member_pathway_reg_subsector_GHG.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reg_subsector_GHG.csv", sep=";", index=False)
        f_logname = _write_data_to_output(df_member_pathway_reg_subsector_GHG, "member_pathway_reg_subsector_GHG", log_name, log_type)

        print("\n5. Merge and sum to Kyoto emissions")
        vars_subsector_pathway = vars_subsector.copy().insert(0, "Emissions|Kyoto Gases")
        # 6. Create member pathways for Kyoto emissions: determine Kyoto emissions covered and not covered by initiatives
        # 6a. determine baseline Kyoto emissions not covered/covered by initiatives

        # remove bunkers for regions
        df_baseline_Kyoto = image_data.read_vars_CO2eq([baseline_scenario], ["Emissions|Kyoto Gases"], dir_RT, GWP, add_calc_variables=False)
        df_baseline_Kyoto.rename(columns={"Scenario": "Scenario_baseline"}, inplace=True)
        df_baseline_Kyoto_covered = image_data.read_vars_CO2eq([baseline_scenario], vars_subsector, dir_RT, GWP, add_calc_variables=False)
        df_baseline_Kyoto_covered.rename(columns={"Scenario": "Scenario_baseline"}, inplace=True)
        mask_bunkers = (df_baseline_Kyoto_covered["Variable"].isin(vars_bunkers)) & ~(df_baseline_Kyoto_covered["Region"]=="World")
        df_baseline_Kyoto_covered = df_baseline_Kyoto_covered[~mask_bunkers]
        df_baseline_Kyoto_covered = df_baseline_Kyoto_covered.groupby(["Region", "Year"], observed=False)["Value"].sum().reset_index()
        f_logname = _write_data_to_output(df_baseline_Kyoto_covered, "baseline_Kyoto_covered", log_name, log_type)
        df_baseline_Kyoto_not_covered = pd.merge(df_baseline_Kyoto, df_baseline_Kyoto_covered, on=["Region", "Year"], how="left", suffixes=("_baseline", "_covered"))
        df_baseline_Kyoto_not_covered["Value"] = df_baseline_Kyoto_not_covered["Value_baseline"] - df_baseline_Kyoto_not_covered["Value_covered"]
        df_baseline_Kyoto_not_covered.drop(["Value_baseline", "Value_covered"], axis=1, inplace=True)
        f_logname = _write_data_to_output(df_baseline_Kyoto_not_covered, "baseline_Kyoto_not_covered", log_name, log_type)
        # 6b. calculate emissions levels member pathways, emissions not covered by initiatives is assumed to follow baseline scenario
        df_member_pathway_Kyoto_emissions = df_member_pathway_reg_subsector_GHG.groupby(["Scenario", "Region", "Year"])["Value"].sum().reset_index()
        df_member_pathway_Kyoto_emissions = pd.merge(df_baseline_Kyoto_not_covered, df_member_pathway_Kyoto_emissions, on=["Region", "Year"], how="left", suffixes=("_not_covered", "_members"))
        df_member_pathway_Kyoto_emissions["Value"] = df_member_pathway_Kyoto_emissions["Value_not_covered"] + df_member_pathway_Kyoto_emissions["Value_members"]
        df_member_pathway_Kyoto_emissions.drop(["Value_not_covered", "Value_members"], axis=1, inplace=True)
        df_member_pathway_Kyoto_emissions.drop(["Scenario_baseline"], axis=1, inplace=True)
        df_member_pathway_Kyoto_emissions["Sector"] = "Total"
        df_member_pathway_Kyoto_emissions["GHG"] = "Total"
        f_logname = _write_data_to_output(df_member_pathway_Kyoto_emissions, "member_pathway_reg_Kyoto", log_name, log_type)
        df_member_pathway_Kyoto_emissions.to_csv(f"{DIR}/data/output/downscale_members/{output_dir}/sector/df_member_pathway_reg_Kyoto.csv", sep=";", index=False)
        df_Kyoto_full = image_data.read_vars_CO2eq([baseline_scenario]+initiative_scenarios+[initiative_scenarios_additional], ["Emissions|Kyoto Gases"], dir_RT, GWP, add_calc_variables=False)
        f_logname = _write_data_to_output(df_Kyoto_full, "Kyoto_full", log_name, log_type)

        return df_overview_initiatives # iniatives and corresponding variables for which data is created in csv files

    except Exception as exc:
        import os
        print(f"\n\033[92mERROR: Something went wrong\033[0m")
        traceback.print_exc()
    finally:
        # save log file
        if not ("df_member_pathway_reg_subsector_GHG" in locals()) or not ("df_member_pathway_Kyoto_emissions" in locals()):
            return pd.DataFrame(), pd.DataFrame()
        else:
            if log_type=="pkl":
                print(f"Saving {log_name}.xlsx to Excel")
                with pd.ExcelWriter(f"{log_name}.xlsx") as writer:
                    for data in log_dataframes:
                        for name, dataframe in data.items():
                            dataframe.to_excel(writer, sheet_name=name)
                return df_member_pathway_reg_subsector_GHG, df_member_pathway_Kyoto_emissions
            elif log_type=="xlsx":
                pass
            else:
                print("Unkown log file type\n")
                return pd.DataFrame(), pd.DataFrame()

def main():
    pass

if __name__ == "__main__":
    main()



